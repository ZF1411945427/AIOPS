
import openpyxl

EXCEL = r"E:\AIOPS\project03\功能测试\AIOPS_功能测试计划_v1.0.xlsx"
wb = openpyxl.load_workbook(EXCEL)

# 看 2-3 个 Sheet 的用例编号格式
for sheet_name in wb.sheetnames[:5]:
    if sheet_name in ["封面", "执行总览"]:
        continue
    ws = wb[sheet_name]
    print("=== %s ===" % sheet_name)
    # 找用例编号列
    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col).value
        if header and "编号" in str(header):
            print("  用例编号列: col=%d header=%s" % (col, header))
            # 打印前5行的编号
            for row in range(2, min(7, ws.max_row+1)):
                val = ws.cell(row=row, column=col).value
                print("    row %d: %s" % (row, val))
            break
    # 找测试结果列
    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col).value
        if header and "测试结果" in str(header):
            print("  测试结果列: col=%d header=%s" % (col, header))
            break
    print()

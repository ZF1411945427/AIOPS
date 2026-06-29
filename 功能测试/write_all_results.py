# -*- coding: utf-8 -*-
"""将全部 E2E 测试结果回写到功能测试计划 Excel"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import json, os

EXCEL = r"E:\AIOPS\project03\功能测试\AIOPS_功能测试计划_v1.0.xlsx"
RESULTS = r"E:\AIOPS\project03\功能测试\all_test_results.json"

# 加载测试结果
with open(RESULTS, encoding="utf-8") as f:
    all_results = json.load(f)

# 构建用例ID -> 结果映射
result_map = {}
for scenario, results in all_results.items():
    for r in results:
        result_map[r["id"]] = r

print(f"共加载 {len(result_map)} 个测试结果")

# 加载 Excel
wb = openpyxl.load_workbook(EXCEL)

# 颜色定义
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
pass_font = Font(color="006100", bold=True)
fail_font = Font(color="9C0006", bold=True)
skip_font = Font(color="9C6500")

# 遍历每个 Sheet
total_updated = 0
for sheet_name in wb.sheetnames:
    if sheet_name in ["封面", "执行总览"]:
        continue
    
    ws = wb[sheet_name]
    
    # 表头在第2行，用例编号在第1列，测试结果在第8列，备注在第9列
    result_col = 8
    remark_col = 9
    case_id_col = 1
    
    # 验证表头
    header_id = ws.cell(row=2, column=1).value
    if not header_id or "编号" not in str(header_id):
        continue
    
    # 遍历每一行（从第3行开始）
    for row in range(3, ws.max_row + 1):
        case_id = ws.cell(row=row, column=case_id_col).value
        if not case_id:
            continue
        
        case_id = str(case_id).strip()
        
        if case_id in result_map:
            r = result_map[case_id]
            status = r["status"]
            detail = r.get("detail", "")
            
            cell = ws.cell(row=row, column=result_col)
            cell.value = status
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
            if status == "PASS":
                cell.fill = green_fill
                cell.font = pass_font
            elif status == "FAIL":
                cell.fill = red_fill
                cell.font = fail_font
            
            if remark_col:
                remark_cell = ws.cell(row=row, column=remark_col)
                remark_cell.value = detail[:200] if detail else ""
                remark_cell.alignment = Alignment(wrap_text=True, vertical="center")
            
            total_updated += 1

# 更新执行总览 Sheet
if "执行总览" in wb.sheetnames:
    ws_overview = wb["执行总览"]
    
    # 找到空行追加统计
    summary_row = ws_overview.max_row + 2
    
    # 计算统计
    total = len(result_map)
    passed = sum(1 for r in result_map.values() if r["status"] == "PASS")
    failed = sum(1 for r in result_map.values() if r["status"] == "FAIL")
    
    ws_overview.cell(row=summary_row, column=1, value="E2E 自动化测试统计 (最终)")
    ws_overview.cell(row=summary_row, column=1).font = Font(bold=True, size=12)
    
    ws_overview.cell(row=summary_row+1, column=1, value="已执行用例")
    ws_overview.cell(row=summary_row+1, column=2, value=total)
    ws_overview.cell(row=summary_row+2, column=1, value="通过")
    ws_overview.cell(row=summary_row+2, column=2, value=passed)
    ws_overview.cell(row=summary_row+2, column=2).fill = green_fill
    ws_overview.cell(row=summary_row+3, column=1, value="失败")
    ws_overview.cell(row=summary_row+3, column=2, value=failed)
    ws_overview.cell(row=summary_row+3, column=2).fill = red_fill
    ws_overview.cell(row=summary_row+4, column=1, value="通过率")
    ws_overview.cell(row=summary_row+4, column=2, value=f"{passed/total*100:.1f}%")
    ws_overview.cell(row=summary_row+4, column=2).font = Font(bold=True)

# 保存
wb.save(EXCEL)
print(f"已更新 {total_updated} 个用例的测试结果到 Excel")
print(f"Excel 已保存: {EXCEL}")

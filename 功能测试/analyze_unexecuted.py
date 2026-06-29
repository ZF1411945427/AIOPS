
import openpyxl, json, os
from collections import Counter

EXCEL = r"E:\AIOPS\project03\功能测试\AIOPS_功能测试计划_v1.0.xlsx"
wb = openpyxl.load_workbook(EXCEL)

unexecuted = []
for sheet_name in wb.sheetnames:
    if sheet_name in ["封面", "测试执行总览", "前置数据需求汇总"]:
        continue
    ws = wb[sheet_name]
    id_col = result_col = precond_col = func_col = None
    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=5):
        for cell in row:
            val = str(cell.value or "")
            if "用例编号" in val or val.strip() == "编号":
                id_col = cell.column
                header_row = cell.row
            if "测试结果" in val:
                result_col = cell.column
            if "前置数据" in val:
                precond_col = cell.column
            if "功能点" in val:
                func_col = cell.column
    if not id_col or not result_col:
        continue
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
        case_id = row[id_col - 1].value if id_col <= len(row) else None
        result = row[result_col - 1].value if result_col <= len(row) else None
        func = row[func_col - 1].value if func_col and func_col <= len(row) else ""
        precond = row[precond_col - 1].value if precond_col and precond_col <= len(row) else ""
        if case_id and result and str(result).strip() == "未执行":
            unexecuted.append({
                "id": str(case_id).strip(),
                "module": sheet_name,
                "func": str(func or ""),
                "precond": str(precond or ""),
            })

print(f"未执行用例: {len(unexecuted)}")
print()
module_count = Counter(u["module"] for u in unexecuted)
print("=== 未执行用例按模块分布 ===")
for mod, cnt in module_count.most_common():
    print(f"  {mod}: {cnt}")

print()
precond_count = Counter()
for u in unexecuted:
    p = u["precond"]
    if p and p not in ("无", "None", ""):
        precond_count[p] += 1
    else:
        precond_count["无前置数据(可直接测)"] += 1
print("=== 未执行用例前置数据需求 ===")
for p, cnt in precond_count.most_common(25):
    print(f"  [{cnt:3}] {p}")

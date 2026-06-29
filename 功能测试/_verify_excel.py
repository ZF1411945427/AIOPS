
import openpyxl
from collections import Counter

EXCEL = r"E:\AIOPS\project03\功能测试\AIOPS_功能测试计划_v1.0.xlsx"
wb = openpyxl.load_workbook(EXCEL)

total = 0
status_count = Counter()
updated = 0

for sheet_name in wb.sheetnames:
    if sheet_name in ["测试计划封面", "前置数据需求汇总", "执行总览"]:
        continue
    ws = wb[sheet_name]
    
    # 表头在第2行，数据从第3行开始
    for row in range(3, ws.max_row + 1):
        case_id = ws.cell(row=row, column=1).value
        if not case_id:
            continue
        result = ws.cell(row=row, column=8).value
        total += 1
        if result:
            status_count[result] += 1
            updated += 1

print("=== Excel 验证 ===")
print("总用例数: %d" % total)
print("已填写结果: %d" % updated)
print("未填写: %d" % (total - updated))
print()
for status, count in sorted(status_count.items()):
    print("  %s: %d" % (status, count))
print()
if updated > 0:
    passed = status_count.get("PASS", 0)
    print("通过率: %.1f%%" % (passed / updated * 100))

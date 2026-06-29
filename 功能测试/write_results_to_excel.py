# -*- coding: utf-8 -*-
"""将 E2E 测试结果回写到功能测试计划 Excel"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import json, os

EXCEL = r"E:\AIOPS\project03\功能测试\AIOPS_功能测试计划_v1.0.xlsx"
RESULTS = r"E:\AIOPS\project03\功能测试\test_results_summary.json"

# 读取测试结果
with open(RESULTS, encoding="utf-8") as f:
    case_map = json.load(f)

print(f"加载测试结果: {len(case_map)} 个用例")

# 打开 Excel
wb = openpyxl.load_workbook(EXCEL)

# 样式
pass_font = Font(color="008000", bold=True)
fail_font = Font(color="FF0000", bold=True)
pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
skip_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
center = Alignment(horizontal="center", vertical="center", wrap_text=True)

updated = 0
not_found = []

# 遍历所有 Sheet（跳过封面和总览）
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    if sheet_name in ["封面", "测试执行总览", "前置数据需求汇总"]:
        continue
    
    # 找到表头行，定位"测试结果"列和"用例编号"列
    header_row = None
    id_col = None
    result_col = None
    detail_col = None
    
    for row in ws.iter_rows(min_row=1, max_row=5, values_only=False):
        for cell in row:
            val = str(cell.value or "")
            if "用例编号" in val or "编号" in val:
                id_col = cell.column
                header_row = cell.row
            if "测试结果" in val or "结果" == val.strip():
                result_col = cell.column
            if "备注" in val or "说明" in val:
                detail_col = cell.column
    
    if not id_col or not result_col:
        continue
    
    if not detail_col:
        detail_col = result_col + 1
    
    # 遍历数据行
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
        id_cell = row[id_col - 1] if id_col <= len(row) else None
        if not id_cell or not id_cell.value:
            continue
        
        case_id = str(id_cell.value).strip()
        
        if case_id in case_map:
            result = case_map[case_id]
            status = result["status"]
            detail = result["detail"][:80] if result["detail"] else ""
            
            # 写入测试结果
            result_cell = ws.cell(row=id_cell.row, column=result_col, value=status)
            result_cell.alignment = center
            
            # 写入备注（失败原因）
            detail_cell = ws.cell(row=id_cell.row, column=detail_col, value=detail)
            detail_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            
            if status == "PASS":
                result_cell.font = pass_font
                result_cell.fill = pass_fill
            elif status == "FAIL":
                result_cell.font = fail_font
                result_cell.fill = fail_fill
                detail_cell.font = fail_font
            
            updated += 1
            del case_map[case_id]  # 已匹配
        else:
            # 未测试的用例标注"未执行"
            result_cell = ws.cell(row=id_cell.row, column=result_col, value="未执行")
            result_cell.alignment = center
            result_cell.fill = skip_fill
            result_cell.font = Font(color="808080")

# 更新测试执行总览 Sheet
if "测试执行总览" in wb.sheetnames:
    ws_overview = wb["测试执行总览"]
    # 找到统计行，更新数据
    total_tested = updated
    total_pass = sum(1 for v in json.load(open(RESULTS, encoding="utf-8")).values() if v["status"] == "PASS")
    total_fail = sum(1 for v in json.load(open(RESULTS, encoding="utf-8")).values() if v["status"] == "FAIL")
    
    # 在总览 Sheet 末尾添加测试结果统计
    stats_row = ws_overview.max_row + 2
    ws_overview.cell(row=stats_row, column=1, value="E2E 测试执行统计").font = Font(bold=True, size=14)
    ws_overview.cell(row=stats_row + 1, column=1, value="已执行用例")
    ws_overview.cell(row=stats_row + 1, column=2, value=total_tested)
    ws_overview.cell(row=stats_row + 2, column=1, value="通过")
    ws_overview.cell(row=stats_row + 2, column=2, value=total_pass)
    ws_overview.cell(row=stats_row + 3, column=1, value="失败")
    ws_overview.cell(row=stats_row + 3, column=2, value=total_fail)
    ws_overview.cell(row=stats_row + 4, column=1, value="通过率")
    ws_overview.cell(row=stats_row + 4, column=2, value=f"{total_pass/total_tested*100:.1f}%")

wb.save(EXCEL)
print(f"\nExcel 已更新: {EXCEL}")
print(f"回写用例: {updated}")
print(f"未匹配用例: {len(case_map)}")
if case_map:
    print(f"未匹配的用例编号: {list(case_map.keys())[:20]}")

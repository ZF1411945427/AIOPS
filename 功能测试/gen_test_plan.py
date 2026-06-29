# -*- coding: utf-8 -*-
"""生成 AIOPS 功能测试计划 Excel (含前置数据需求标注)"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os, json

wb = openpyxl.Workbook()

title_font = Font(name='微软雅黑', size=16, bold=True, color='FFFFFF')
title_fill = PatternFill('solid', fgColor='1F2937')
module_font = Font(name='微软雅黑', size=13, bold=True, color='FFFFFF')
module_fill = PatternFill('solid', fgColor='4F46E5')
header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
header_fill = PatternFill('solid', fgColor='6366F1')
normal_font = Font(name='微软雅黑', size=10)
bold_font = Font(name='微软雅黑', size=11, bold=True)
high_fill = PatternFill('solid', fgColor='FEE2E2')
medium_fill = PatternFill('solid', fgColor='FEF3C7')
low_fill = PatternFill('solid', fgColor='D1FAE5')
total_fill = PatternFill('solid', fgColor='E0E7FF')
need_data_fill = PatternFill('solid', fgColor='FEE2E2')
no_data_fill = PatternFill('solid', fgColor='D1FAE5')
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
left = Alignment(horizontal='left', vertical='center', wrap_text=True)
thin_border = Border(
    left=Side(style='thin', color='D1D5DB'),
    right=Side(style='thin', color='D1D5DB'),
    top=Side(style='thin', color='D1D5DB'),
    bottom=Side(style='thin', color='D1D5DB')
)

data_path = os.path.join(os.path.dirname(__file__), 'modules_data.json')
with open(data_path, encoding='utf-8') as f:
    modules = json.load(f)

# ========== Sheet 1: 封面 ==========
ws_cover = wb.active
ws_cover.title = '测试计划封面'
ws_cover.sheet_properties.tabColor = '1F2937'
ws_cover.merge_cells('A1:H1')
ws_cover['A1'] = 'AIOPS 智能运维平台 — 功能测试计划 (REAL 模式)'
ws_cover['A1'].font = title_font
ws_cover['A1'].fill = title_fill
ws_cover['A1'].alignment = center
ws_cover.row_dimensions[1].height = 50

cover_info = [
    ('项目名称', 'AIOPS 智能运维平台'),
    ('测试环境', 'REAL 模式 (aiops_real.db) — 真实测试数据'),
    ('DEMO 模式', '仅作模拟展示，不用于功能测试'),
    ('后端地址', 'http://127.0.0.1:8000'),
    ('前端地址', 'http://127.0.0.1:3000 (Vue SPA)'),
    ('技术栈', 'Python 3.13 + FastAPI + Vue 3 + Element Plus + SQLite'),
    ('API 端点总数', '304 个 (79 个路由模块)'),
    ('测试方式', 'Playwright E2E 自动化 + API 接口验证 + UI 交互验证'),
    ('测试账号', 'admin / admin123'),
    ('生成日期', '2026-06-28'),
    ('版本', 'v1.1 (增加前置数据标注)'),
]
row = 3
for label, value in cover_info:
    c = ws_cover.cell(row=row, column=1, value=label)
    c.font = bold_font
    c.fill = PatternFill('solid', fgColor='F3F4F6')
    c.alignment = center
    c.border = thin_border
    ws_cover.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
    c2 = ws_cover.cell(row=row, column=2, value=value)
    c2.font = normal_font
    c2.alignment = left
    c2.border = thin_border
    row += 1

# 模块统计
row += 2
ws_cover.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
c = ws_cover.cell(row=row, column=1, value='测试模块统计')
c.font = module_font
c.fill = module_fill
c.alignment = center
row += 1

stats_header = ['序号', '一级模块', '用例数', '高', '中', '低', '自动化', '需前置数据']
for col, h in enumerate(stats_header, 1):
    c = ws_cover.cell(row=row, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center
    c.border = thin_border
stats_start = row

# ========== 前置数据汇总 Sheet ==========
ws_precond = wb.create_sheet('前置数据需求汇总')
ws_precond.sheet_properties.tabColor = 'F59E0B'
ws_precond.merge_cells('A1:E1')
ws_precond['A1'] = '前置数据需求汇总 — REAL 模式'
ws_precond['A1'].font = title_font
ws_precond['A1'].fill = title_fill
ws_precond['A1'].alignment = center
ws_precond.row_dimensions[1].height = 40

precond_header = ['数据类型', '需求数量', '涉及用例示例', '获取方式', '备注']
for col, h in enumerate(precond_header, 1):
    c = ws_precond.cell(row=2, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center
    c.border = thin_border

# ========== 创建模块 Sheet ==========
total_cases = 0
high_count = 0
medium_count = 0
low_count = 0
auto_count = 0
need_data_count = 0
module_stats = []

# 收集所有前置数据类型
precond_types = {}

for mod_name, color, cases in modules:
    ws = wb.create_sheet(title=mod_name[:31])
    ws.sheet_properties.tabColor = color
    ws.merge_cells('A1:H1')
    ws['A1'] = mod_name
    ws['A1'].font = module_font
    ws['A1'].fill = module_fill
    ws['A1'].alignment = center
    ws.row_dimensions[1].height = 35

    headers = ['用例编号', '功能点', '测试步骤', '预期结果', '优先级', '测试类型', '前置数据需求', '测试结果']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border
    ws.row_dimensions[2].height = 30

    h = m = l = a = nd = 0
    for i, case_data in enumerate(cases):
        case_id, func, steps, expected, priority, test_type, precond = case_data
        r = i + 3
        data = [case_id, func, steps, expected, priority, test_type, precond, '']
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.font = normal_font
            cell.alignment = left if col in (2, 3, 4, 7) else center
            cell.border = thin_border
        if priority == '高':
            ws.cell(row=r, column=5).fill = high_fill
            h += 1
        elif priority == '中':
            ws.cell(row=r, column=5).fill = medium_fill
            m += 1
        else:
            ws.cell(row=r, column=5).fill = low_fill
            l += 1
        if 'E2E' in test_type:
            a += 1
        # 前置数据颜色标注
        if '无需' in precond:
            ws.cell(row=r, column=7).fill = no_data_fill
        else:
            ws.cell(row=r, column=7).fill = need_data_fill
            nd += 1
            need_data_count += 1
            # 收集数据类型
            if precond not in precond_types:
                precond_types[precond] = []
            precond_types[precond].append(case_id)

    total_cases += len(cases)
    high_count += h
    medium_count += m
    low_count += l
    auto_count += a
    module_stats.append((mod_name, len(cases), h, m, l, a, nd))

    widths = [12, 22, 42, 42, 8, 12, 32, 12]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

# 填充封面统计
for idx, (mod_name, count, h, m, l, a, nd) in enumerate(module_stats, 1):
    r = stats_start + idx
    vals = [idx, mod_name, count, h, m, l, a, nd]
    for col, val in enumerate(vals, 1):
        cell = ws_cover.cell(row=r, column=col, value=val)
        cell.font = normal_font
        cell.alignment = center
        cell.border = thin_border

r = stats_start + len(module_stats) + 1
vals = ['合计', '', total_cases, high_count, medium_count, low_count, auto_count, need_data_count]
for col, val in enumerate(vals, 1):
    cell = ws_cover.cell(row=r, column=col, value=val)
    cell.font = bold_font
    cell.fill = total_fill
    cell.alignment = center
    cell.border = thin_border

for col, w in enumerate([10, 22, 10, 8, 8, 8, 10, 14], 1):
    ws_cover.column_dimensions[get_column_letter(col)].width = w

# 填充前置数据汇总
precond_info = [
    ('真实资产(SSH可连)', '需1台可SSH的服务器，用于资产CRUD+测试连接+脚本执行+指标采集', 'ASSET-004/007, AUTO-007'),
    ('真实资产(DB已有)', 'REAL库需有资产记录，用于列表/搜索/编辑/删除', 'ASSET-001/002/005/006'),
    ('指标数据', '需通过采集器对online资产采集指标，写入metric_records表', 'METRIC-001~010, PRED-001~007'),
    ('告警数据', '需触发告警规则产生告警记录，或手动插入测试告警', 'ALERT-001~015, DASH-004/005/009'),
    ('告警规则', '需在REAL库创建告警规则(指标+阈值)', 'ALERT-006~010'),
    ('事件数据', '需有incident记录，可通过事件源接入或手动创建', 'INC-001~004, EVENT-001/002'),
    ('Trace数据', '需通过OTLP/Jaeger接入链路数据', 'TRACE-001~004/007'),
    ('日志数据', '需接入日志数据(ES/Filebeat等)', 'LOG-001~004, INC-006'),
    ('拓扑数据', '需创建资产间拓扑关系', 'TOPO-001~005'),
    ('K8s集群', '需可访问的K8s集群(kubeconfig配置)', 'K8S-001~019'),
    ('Docker主机', '需可访问的Docker引擎', 'K8S-017/018'),
    ('AI配置', '需AI提供商配置(已配好aicodee-MiniMax)', 'AI-001~015'),
    ('AI会话数据', '需有历史AI对话会话', 'AI-005/006/016'),
    ('知识库数据', '需创建知识库记录', 'KB-001~005'),
    ('通知渠道', '需配置通知渠道(邮件/钉钉/企业微信)', 'NOTIF-001~006'),
    ('自愈规则/工作流', '需创建自愈规则和工作流', 'AUTO-001~006'),
    ('变更数据', '需创建变更工单', 'AUTO-010/011/012'),
    ('异常检测配置', '需创建异常检测配置', 'ANOM-001~005'),
    ('预测模型', '需创建预测模型', 'PRED-005/006'),
    ('ES服务', '需可访问的Elasticsearch', 'SYS-013'),
    ('Kafka服务', '需可访问的Kafka', 'SYS-014'),
    ('外部数据源', '需可访问的外部数据源(Prometheus等)', 'SYS-018'),
    ('NetFlow采集器', '需NetFlow采集器', 'SYS-019'),
    ('外部Webhook', '需可接收POST的外部URL', 'ALERT-013'),
    ('外部CMDB', '需外部CMDB服务地址+凭据', 'ASSET-013'),
    ('外部事件源', '需可访问的事件源服务', 'EVENT-003/004'),
    ('外部通知服务', '需邮件/钉钉/企业微信服务', 'NOTIF-002'),
    ('外部服务网格', '需可访问的服务网格(如Istio)', 'TOPO-006'),
    ('外部Trace Agent', '需可发送OTLP数据的Agent', 'TRACE-006'),
    ('Runbook数据', '需创建Runbook', 'AUTO-013'),
]

for idx, (dtype, method, examples) in enumerate(precond_info, 1):
    r = idx + 2
    vals = [dtype, '', examples, method, '']
    # count
    count = 0
    for precond, case_ids in precond_types.items():
        if dtype.split('(')[0] in precond:
            count += len(case_ids)
    vals[1] = count
    for col, val in enumerate(vals, 1):
        cell = ws_precond.cell(row=r, column=col, value=val)
        cell.font = normal_font
        cell.alignment = left if col in (3, 4) else center
        cell.border = thin_border

for col, w in enumerate([20, 10, 35, 40, 15], 1):
    ws_precond.column_dimensions[get_column_letter(col)].width = w

# 测试执行总览
ws_exec = wb.create_sheet('测试执行总览')
ws_exec.sheet_properties.tabColor = '059669'
ws_exec.merge_cells('A1:G1')
ws_exec['A1'] = '测试执行总览'
ws_exec['A1'].font = title_font
ws_exec['A1'].fill = title_fill
ws_exec['A1'].alignment = center
ws_exec.row_dimensions[1].height = 40

exec_headers = ['模块', '用例总数', '已执行', '通过', '失败', '阻塞', '通过率']
for col, h in enumerate(exec_headers, 1):
    cell = ws_exec.cell(row=2, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = thin_border

for idx, (mod_name, count, h, m, l, a, nd) in enumerate(module_stats, 1):
    r = idx + 2
    vals = [mod_name, count, 0, 0, 0, 0, '0%']
    for col, val in enumerate(vals, 1):
        cell = ws_exec.cell(row=r, column=col, value=val)
        cell.font = normal_font
        cell.alignment = center
        cell.border = thin_border

r = len(module_stats) + 3
vals = ['合计', total_cases, 0, 0, 0, 0, '0%']
for col, val in enumerate(vals, 1):
    cell = ws_exec.cell(row=r, column=col, value=val)
    cell.font = bold_font
    cell.fill = total_fill
    cell.alignment = center
    cell.border = thin_border

for col, w in enumerate([25, 12, 12, 12, 12, 12, 12], 1):
    ws_exec.column_dimensions[get_column_letter(col)].width = w

# 保存
output_path = os.path.join(os.path.dirname(__file__), 'AIOPS_功能测试计划_v1.0.xlsx')
wb.save(output_path)
print(f"Excel saved: {output_path}")
print(f"Total cases: {total_cases}")
print(f"High: {high_count}, Medium: {medium_count}, Low: {low_count}")
print(f"Auto cases: {auto_count}")
print(f"Need precond data: {need_data_count}")
print(f"No precond needed: {total_cases - need_data_count}")
print(f"Modules: {len(modules)}")
print(f"Sheets: {len(wb.sheetnames)}")

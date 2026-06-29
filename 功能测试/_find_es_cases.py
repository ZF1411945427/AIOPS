
import openpyxl

EXCEL = r"E:\AIOPS\project03\功能测试\AIOPS_功能测试计划_v1.0.xlsx"
wb = openpyxl.load_workbook(EXCEL)

# 找所有 ES/日志/Trace 相关的用例
keywords = ["ES", "elastic", "日志", "log", "trace", "链路", "Trace", "ELK"]

for sheet_name in wb.sheetnames:
    if sheet_name in ["测试计划封面", "前置数据需求汇总", "执行总览"]:
        continue
    ws = wb[sheet_name]
    for row in range(3, ws.max_row + 1):
        case_id = ws.cell(row=row, column=1).value
        if not case_id:
            continue
        func_point = ws.cell(row=row, column=2).value or ""
        test_step = ws.cell(row=row, column=3).value or ""
        expected = ws.cell(row=row, column=4).value or ""
        precondition = ws.cell(row=row, column=7).value or ""
        result = ws.cell(row=row, column=8).value or ""
        remark = ws.cell(row=row, column=9).value or ""
        
        all_text = str(case_id) + str(func_point) + str(test_step) + str(expected) + str(precondition)
        if any(kw in all_text for kw in keywords):
            print("=== %s | %s ===" % (sheet_name, case_id))
            print("  功能点: %s" % str(func_point)[:80])
            print("  测试步骤: %s" % str(test_step)[:100])
            print("  预期结果: %s" % str(expected)[:100])
            print("  前置数据: %s" % str(precondition)[:60])
            print("  测试结果: %s" % str(result))
            print("  备注: %s" % str(remark)[:80])
            print()

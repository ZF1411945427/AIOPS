
import openpyxl

EXCEL = r"E:\AIOPS\project03\功能测试\AIOPS_功能测试计划_v1.0.xlsx"
wb = openpyxl.load_workbook(EXCEL)

# 看认证与登录 Sheet 的所有表头
ws = wb["认证与登录"]
print("=== 认证与登录 表头 ===")
for col in range(1, ws.max_column + 1):
    val = ws.cell(row=1, column=col).value
    print("  col %d: %s" % (col, val))

print("\n=== 认证与登录 前3行数据 ===")
for row in range(2, 5):
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=row, column=col).value
        if val:
            print("  row %d col %d: %s" % (row, col, str(val)[:60]))

# 看资产管理
ws2 = wb["资产管理 (CMDB)"]
print("\n=== 资产管理 表头 ===")
for col in range(1, ws2.max_column + 1):
    val = ws2.cell(row=1, column=col).value
    print("  col %d: %s" % (col, val))
